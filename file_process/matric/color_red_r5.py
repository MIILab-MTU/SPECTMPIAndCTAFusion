import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
import os

input_file = r"eval_results_bcpdpp.csv"
output_excel = r"eval_results_r5colorcpdpp.xlsx"
numeric_columns = [
    'icp_mse', 'sicp_mse', 'rigid_mse', 'affine_mse',
    'icp_juli', 'sicp_juli', 'rigid_juli', 'affine_juli',
    'icp_jl', 'sicp_jl', 'rigid_jl', 'affine_jl'
]

def main():
    output_dir = os.path.dirname(output_excel)
    os.makedirs(output_dir, exist_ok=True)
    try:
        df = pd.read_csv(input_file)
    except FileNotFoundError:
        print(f"Error: Input file {input_file} does not exist!")
        return
    except Exception as e:
        print(f"Error reading {input_file}: {str(e)}")
        return
    missing_columns = [col for col in numeric_columns if col not in df.columns]
    if missing_columns:
        print(f"Error: Missing columns: {missing_columns}")
        return
    print("Original data (first 2 rows):")
    print(df[numeric_columns].head(2))
    for col in numeric_columns:
        df[col] = df[col].apply(lambda x: round(x, 5) if pd.notnull(x) else x)
    print("\nFormatted data (first 2 rows):")
    print(df[numeric_columns].head(2))
    try:
        df.to_excel(output_excel, index=False, engine='openpyxl')
    except Exception as e:
        print(f"Error saving {output_excel}: {str(e)}")
        return
    try:
        wb = load_workbook(output_excel)
        ws = wb.active
        header = df.columns
        ori_col = header.get_loc('ori') + 1
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=ori_col).value == 'auto':
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill
        wb.save(output_excel)
        print(f"\nFormatted and styled Excel file saved to {output_excel}")
    except Exception as e:
        print(f"Error saving {output_excel}: {str(e)}")

if __name__ == '__main__':
    main()